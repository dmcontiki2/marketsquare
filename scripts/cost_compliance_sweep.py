#!/usr/bin/env python3
"""
cost_compliance_sweep.py — nightly $0 static cost-compliance sweep
===================================================================
Design principles audited (David, 11 Jun 2026):
  P1  $0-first: no AI tokens / paid API calls during normal FREE use;
      Haiku (or cheaper) when unavoidable; paid surfaces metered.
  P2  Cost budgeting: every AI call wrapped in ceiling-check + spend-logging;
      cost workbook kept in step with code pricing ("Cost model impact").
  P3  Independence: paid providers OFF until contracted; know every external
      paid call site (the Google-API scraping incident must never recur).

The sweep itself uses NO AI and NO network (unless MS_BEA_URL+MS_API_KEY are
set, in which case it pulls live spend from /admin/ai-spend/summary).

Usage:  python scripts/cost_compliance_sweep.py [--root <Projects dir>] [--quiet]
Output: Records/COST_SWEEP_<date>.md  + stdout summary.
Exit:   0 clean · 1 warnings · 2 critical findings.
"""
from __future__ import annotations
import argparse, ast, datetime, json, os, re, sys
from pathlib import Path

# ── what counts as a paid call site ──────────────────────────────────────────
PAID_PATTERNS = {
    "Anthropic API":      re.compile(r"api\.anthropic\.com|anthropic-version"),
    "Anthropic SDK":      re.compile(r"^\s*(import anthropic|from anthropic)", re.M),
    "Google APIs":        re.compile(r"(?<!fonts\.)googleapis\.com|GOOGLE_\w*API_KEY|maps\.googleapis"),
    "OpenAI":             re.compile(r"api\.openai\.com|OPENAI_API_KEY"),
    "Paystack (txn)":     re.compile(r"api\.paystack\.co"),
    "Paid data feeds":    re.compile(r"pricecharting\.com/api|lightstone|rentcast\.io|attomdata|corelogic", re.I),
}
MODEL_RE = re.compile(r"claude-([a-z]+)[-0-9a-z.]*")
# "claude-*" tokens that are NOT model families (tools/plugins/filenames that merely start
# with "claude-"). Skipped in model_discipline() so they don't raise spurious "unknown model
# family" warnings. e.g. claude-mem (the memory plugin) and its claude-mem.db data file.
KNOWN_NON_MODELS = {"mem"}
SKIP_DIRS = {"node_modules", "__pycache__", ".git", "archive", "_CCP_STAGED",
             ".ruff_cache", ".claude", "Kronberg", "Obsidian", "Records"}
SKIP_FILE = re.compile(r"\.bak[-.]|\.bak$|~\$|cost_compliance_sweep\.py$")
SCAN_EXT = {".py", ".js", ".html", ".json", ".sh", ".bat"}
# Primary app repo = the parent of this script's dir (.../<repo>/scripts/). Derive its
# name from the filesystem so a brand/folder rename (MarketSquare -> TrustSquare) cannot
# desync the sweep from the directory it must scan and write into.
MAIN_REPO = Path(__file__).resolve().parents[1].name
REPOS = [MAIN_REPO, "AdvertAgent", "CityLauncher", "SellerScraper",
         "Auction", "LaunchCadence", "Codices"]

CRIT, WARN, INFO, OK = "CRITICAL", "WARN", "INFO", "OK"

def iter_files(root: Path):
    for repo in REPOS:
        base = root / repo
        if not base.exists():
            continue
        for p in base.rglob("*"):
            if p.is_dir() or p.suffix.lower() not in SCAN_EXT:
                continue
            if any(d in p.parts for d in SKIP_DIRS) or SKIP_FILE.search(p.name):
                continue
            yield p

def read(p: Path) -> str:
    try:
        return p.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""

# ── 1 · paid call-site inventory ─────────────────────────────────────────────
def inventory(root: Path):
    hits = []
    for p in iter_files(root):
        text = read(p)
        for label, rx in PAID_PATTERNS.items():
            for m in rx.finditer(text):
                line = text.count("\n", 0, m.start()) + 1
                hits.append((label, str(p.relative_to(root)), line))
    return hits

# ── 2 · wrapper compliance: every Anthropic call wrapped (P2) ────────────────
def wrapper_compliance(root: Path):
    findings = []
    targets = [root / MAIN_REPO / "bea_main.py",
               root / "AdvertAgent" / "service" / "advert_agent.py"]
    for path in targets:
        if not path.exists():
            continue
        src = read(path)
        try:
            tree = ast.parse(src)
        except SyntaxError as e:
            findings.append((CRIT, f"{path.name}: does not parse ({e})"))
            continue
        lines = src.splitlines()
        for node in ast.walk(tree):
            if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                continue
            seg = "\n".join(lines[node.lineno - 1: node.end_lineno])
            # The inference URL/headers now live in single-source helpers
            # (_ts_ai_url / _ts_ai_headers, added Jun 2026). An endpoint that
            # POSTs to _ts_ai_url() IS an Anthropic call site even though the
            # literal "api.anthropic.com" no longer appears in its body.
            INDIRECT = ("_ts_ai_url(" in seg) or ("_ts_ai_headers(" in seg)
            DIRECT   = ("api.anthropic.com" in seg) or ("ANTHROPIC_URL" in seg)
            # ...but the helpers themselves are resolvers, not call sites:
            # skip their own defs so the URL dict isn't mis-flagged as unmetered.
            if node.name in ("_ts_ai_url", "_ts_ai_headers", "_ts_active_provider",
                             "_ts_models_for"):
                continue
            if not (DIRECT or INDIRECT):
                continue
            ceil = "_check_cost_ceiling" in seg
            log  = "_log_ai_spend" in seg
            metered = ("bea_commit" in seg or "ai-commit" in seg or
                       "hold" in node.name or "tuppence" in seg.lower())
            if path.name == "advert_agent.py":
                # AdvertAgent paid runs are metered via Tuppence hold/settle
                if metered or "_worker" in node.name or "run" in node.name:
                    findings.append((OK, f"advert_agent.py:{node.lineno} `{node.name}` — metered via Tuppence hold/settle"))
                else:
                    findings.append((WARN, f"advert_agent.py:{node.lineno} `{node.name}` calls Anthropic without visible metering"))
            else:
                tup = "_deduct_tuppence" in seg or "tuppence" in seg.lower()
                helper = node.name.startswith("_") and ("return" in seg and ("_oin" in seg or "tokens" in seg))
                caller_logs = helper and f"{node.name}(" in src.replace(seg, "") and "_log_ai_spend" in src
                if ceil and log:
                    findings.append((OK, f"bea_main.py:{node.lineno} `{node.name}` — ceiling ✓ spend-log ✓"))
                elif caller_logs:
                    findings.append((WARN, f"bea_main.py:{node.lineno} `{node.name}` — helper; caller logs spend, but add a ceiling check"))
                elif tup and not log:
                    findings.append((WARN, f"bea_main.py:{node.lineno} `{node.name}` — Tuppence-metered (revenue covers it) but NOT spend-logged: tokens invisible to the dashboard"))
                elif log and not ceil:
                    findings.append((WARN, f"bea_main.py:{node.lineno} `{node.name}` — spend-log ✓ but NO _check_cost_ceiling"))
                elif ceil and not log:
                    findings.append((WARN, f"bea_main.py:{node.lineno} `{node.name}` — ceiling ✓ but NO _log_ai_spend"))
                else:
                    findings.append((CRIT, f"bea_main.py:{node.lineno} `{node.name}` — UNWRAPPED & UNMETERED Anthropic call (no ceiling, no spend log, no Tuppence)"))
    return findings

# ── 3 · model discipline: Haiku unless paid+metered (P1) ─────────────────────
def model_discipline(root: Path):
    findings = []
    for p in iter_files(root):
        text = read(p)
        tlines = text.splitlines()
        for m in MODEL_RE.finditer(text):
            fam = m.group(1)
            line = text.count("\n", 0, m.start()) + 1
            rel = str(p.relative_to(root))
            if fam in KNOWN_NON_MODELS:
                continue   # e.g. claude-mem / claude-mem.db — a plugin name, not a model
            if fam == "haiku":
                continue
            ltxt = tlines[line - 1].lstrip() if line <= len(tlines) else ""
            # comments / docs and price-table rows are not call sites
            if ltxt.startswith(("#", "//", "*", "<!--")):
                continue
            col = m.start() - (text.rfind("\n", 0, m.start()) + 1)
            if "#" in ltxt and ltxt.index("#") < max(0, col - (len(tlines[line - 1]) - len(ltxt))):
                continue   # match sits inside a trailing comment
            if re.match(r"[A-Z_]*MODEL\s*=", ltxt):
                findings.append((INFO, f"{rel}:{line} model constant `{ltxt.split('=')[0].strip()}` = {m.group(0)} — used by Tuppence-metered endpoints; keep justified"))
                continue
            # Provider-model registry rows map an abstract tier -> a vendor model
            # string, e.g.  "anthropic": {"haiku":..,"sonnet":"claude-sonnet-4-6"}.
            # This is the single-source TASK_MODEL table (ai_provider.py) plus the
            # in-process fallback dict in bea_main.py — the metered registry, not a
            # loose call site. Treat a tier-keyed Sonnet entry as justified.
            if re.search(r'"(?:sonnet|vision|reason|triage)"\s*:', ltxt):
                findings.append((INFO, f"{rel}:{line} Sonnet in the provider-model registry (TASK_MODEL/fallback) — single-source, Tuppence-metered; keep justified"))
                continue
            if re.search(r"\(\s*\d+\.\d+\s*,\s*\d+\.\d+\s*\)", ltxt) or \
               re.search(r"_AI_COST|_MODEL_PRICE|MODEL_RATES|AA_MODEL_FALLBACK", ltxt):
                continue
            if fam == "opus":
                findings.append((CRIT, f"{rel}:{line} uses OPUS ({m.group(0)}) — cost model rejected Opus"))
            elif fam == "sonnet":
                if "advert_agent.py" in rel:
                    findings.append((INFO, f"{rel}:{line} Sonnet — allowed: paid Level-2, Tuppence-metered"))
                else:
                    findings.append((WARN, f"{rel}:{line} Sonnet outside the metered AdvertAgent registry — justify or downgrade to Haiku"))
            else:
                findings.append((WARN, f"{rel}:{line} unknown model family `{m.group(0)}` — classify"))
    return findings

# ── 4 · paid provider flags must be OFF until contracted (P3) ────────────────
def provider_flags(root: Path):
    findings = []
    tiers = root / MAIN_REPO / "ai_service_tiers.py"
    if tiers.exists():
        m = re.search(r"DEFAULT_PROVIDERS\s*:\s*dict\[str,\s*bool\]\s*=\s*\{(.*?)\n\}", read(tiers), re.S)
        if m:
            paid_known = {"pricecharting","pcgs","gocollect","watchcharts","propertydata",
                          "rentcast","attom","loom","lightstone","domain","corelogic",
                          "sqm","caphpi","redbook","transunion_auto"}
            for name, val in re.findall(r'"(\w+)":\s*(True|False)', m.group(1)):
                if name in paid_known and val == "True":
                    findings.append((CRIT, f"ai_service_tiers.py: paid provider `{name}` is ON — contract + B7 ceiling + CHANGELOG line required"))
            if not findings:
                findings.append((OK, "ai_service_tiers.py: all paid providers OFF"))
    ff = root / MAIN_REPO / "feature_flags.json"
    if ff.exists():
        try:
            flags = json.loads(read(ff))
            on = []
            if flags.get("paid_tiers_enabled"):
                on.append("paid_tiers_enabled")
            prov = flags.get("providers") or {}
            paid_set = {"pricecharting","pcgs","gocollect","watchcharts","propertydata",
                        "rentcast","attom","loom","lightstone","domain","corelogic",
                        "sqm","caphpi","redbook","transunion_auto"}
            if isinstance(prov, dict):
                on += [f"providers.{k}" for k, v in prov.items() if v and k in paid_set]
            findings.append((WARN, f"feature_flags.json paid/provider flags ON: {on}") if on
                            else (OK, "feature_flags.json: paid_tiers_enabled=false, all provider flags off"))
        except Exception as e:
            findings.append((WARN, f"feature_flags.json unreadable: {e}"))
    return findings

# ── 5 · BEFORE-YOU-TEST cost surfaces (the repeated-test-runs guard) ─────────
def test_surfaces(root: Path):
    out = []
    ms_html = read(root / MAIN_REPO / "marketsquare.html")
    if 'id="ai-dryrun"' in ms_html:
        checked = bool(re.search(r'id="ai-dryrun"[^>]*checked', ms_html))
        out.append((OK if checked else CRIT,
                    f"AI dry-run toggle default: {'ON — replays fixtures, $0' if checked else 'OFF — every run bills tokens + Tuppence!'}"))
    gm = root / "CityLauncher" / "scraper" / "sources" / "google_maps.py"
    if gm.exists():
        # COST risk only when a LIVE key is actually present (Places API bills per call).
        # Unset or commented => $0 Playwright fallback, i.e. the incident path is closed.
        # Static + $0: check the process env and any committed .env under CityLauncher.
        key_re = re.compile(r"^[ \t]*GOOGLE_MAPS_API_KEY[ \t]*=[ \t]*(\S+)", re.M)
        live = "process env" if os.getenv("GOOGLE_MAPS_API_KEY") else ""
        if not live:
            for envf in (root / "CityLauncher").rglob(".env*"):
                if any(d in envf.parts for d in SKIP_DIRS):
                    continue
                mk = key_re.search(read(envf))
                if mk and mk.group(1).strip(' "\'') not in ("", "''", '""'):
                    live = str(envf.relative_to(root))
                    break
        if live:
            out.append((WARN, f"CityLauncher google_maps.py: GOOGLE_MAPS_API_KEY is SET ({live}) — paid Places "
                              "API will BILL. Unset it for test runs; the Playwright fallback is $0 (the incident path)."))
        else:
            out.append((OK, "CityLauncher google_maps.py: GOOGLE_MAPS_API_KEY unset — $0 Playwright fallback "
                            "active; paid Places API not reachable (the incident path is closed)."))
    out.append((INFO, "Cost-bearing surfaces for live testing: /ai/run (Tuppence + Sonnet tokens), "
                      "/advert-agent/market-note (Haiku), /listings/draft-from-photo (Haiku, template-fallback), "
                      "/listings/photo orientation (Haiku vision), Paystack init (test keys = $0)"))
    out.append((INFO, "Qualifying rule: ONE paid live run per feature; scenario testing uses dry-run "
                      "fixtures or unset ANTHROPIC_API_KEY (all flows fail open to $0 paths)"))
    out.append((INFO, "Ceilings live in DB (ai_spend_config); ceiling 0 = OFF. Before any test day set "
                      "daily_user/platform ceilings low (e.g. $1/$5) and verify via /admin/ai-spend/summary. "
                      "Authoritative check: live_spend() flags CRITICAL if the platform ceiling is 0/unset when "
                      "MS_BEA_URL+MS_API_KEY are set — a static $0 sweep cannot read the DB itself."))
    return out

# ── 6 · cost-workbook drift (P2) ─────────────────────────────────────────────
def workbook_drift(root: Path):
    findings = []
    ms = root / MAIN_REPO
    wb_path = ms / "Cost_Breakdown_GlobalLaunch.xlsx"
    bea = read(ms / "bea_main.py")
    code_tiers = dict(re.findall(r'"(\w+)":\s*\{"amount_rands[^}]*"usd":\s*(\d+)', bea))
    if not wb_path.exists():
        return [(WARN, "Cost workbook not found")]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        labels = []
        if "Assumptions" in wb.sheetnames:
            for row in wb["Assumptions"].iter_rows(max_col=2, values_only=True):
                a, b = (row + (None, None))[:2]
                if a and "price/month" in str(a).lower():
                    labels.append((str(a).strip(), b))
        wb_prices = {re.sub(r"\s*price/month.*", "", l).strip().lower(): v for l, v in labels}
        new_model = {"starter": 5, "pro": 20, "free": 0, "agency": 0}
        stale = [l for l in wb_prices if l in ("standard", "professional", "business", "elite")]
        missing = [t for t in ("starter", "pro") if t not in wb_prices]
        if stale or missing:
            findings.append((WARN, f"Workbook Assumptions still on OLD tiers {stale}; Simpler Model tiers missing {missing} "
                                   f"— code says {[(k, v) for k, v in code_tiers.items() if k in new_model]}. Update the workbook."))
        else:
            findings.append((OK, "Workbook tier assumptions match the Simpler Model"))
    except ImportError:
        findings.append((INFO, "openpyxl not installed — workbook content check skipped (pip install openpyxl)"))
    except Exception as e:
        findings.append((WARN, f"Workbook check failed: {e}"))
    try:
        wb_m = datetime.date.fromtimestamp(wb_path.stat().st_mtime)
        chl = read(ms / "CHANGELOG.md")
        mc = re.search(r"## Session [\w·\s]*?(\d{1,2} \w+ 20\d\d).{0,4000}?Cost model impact", chl, re.S)
        findings.append((INFO, f"Workbook last modified {wb_m}; latest CHANGELOG cost-impact entry: "
                               f"{mc.group(1) if mc else 'not found'} — reconcile if the code moved later"))
    except Exception:
        pass
    return findings

# ── 7 · live spend (optional, via admin endpoint) ────────────────────────────
def live_spend():
    url, key = os.getenv("MS_BEA_URL"), os.getenv("MS_API_KEY")
    if not (url and key):
        return [(INFO, "Live spend: set MS_BEA_URL + MS_API_KEY to pull /admin/ai-spend/summary (endpoint staged 11 Jun)")]
    try:
        import urllib.request
        req = urllib.request.Request(url.rstrip("/") + "/admin/ai-spend/summary",
                                     headers={"X-Api-Key": key})
        data = json.loads(urllib.request.urlopen(req, timeout=15).read())
        out = [(INFO, f"Live spend today ${data.get('today_usd', 0):.2f} · 7-day ${data.get('week_usd', 0):.2f} · "
                      f"ceilings user=${data.get('daily_user_ceiling_usd', 0)} platform=${data.get('daily_platform_ceiling_usd', 0)}")]
        if not data.get("daily_platform_ceiling_usd"):
            out.append((CRIT, "Platform daily ceiling is 0/unset — AI spend is UNCAPPED on the live server"))
        for row in (data.get("by_endpoint") or [])[:8]:
            out.append((INFO, f"  {row['endpoint']}: ${row['usd']:.2f} ({row['calls']} calls, 7d)"))
        return out
    except Exception as e:
        return [(WARN, f"Live spend pull failed: {e}")]

# ── report ───────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=None)
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()
    root = Path(args.root) if args.root else Path(__file__).resolve().parents[2]
    today = datetime.date.today().isoformat()

    sections = [
        ("Wrapper compliance — every AI call ceiling-checked + spend-logged (P2)", wrapper_compliance(root)),
        ("Model discipline — Haiku unless paid + metered (P1)", model_discipline(root)),
        ("Paid provider flags — OFF until contracted (P3)", provider_flags(root)),
        ("BEFORE YOU TEST — live-cost surfaces & guards", test_surfaces(root)),
        ("Cost-workbook drift (P2)", workbook_drift(root)),
        ("Live spend", live_spend()),
    ]
    inv = inventory(root)

    counts = {CRIT: 0, WARN: 0, INFO: 0, OK: 0}
    lines = [f"# Cost-Compliance Sweep — {today}",
             f"_Principles: P1 $0-first · P2 budget every call · P3 independence/hot-swap. "
             f"Sweep is static + $0; scanned {len(REPOS)} repos under `{root}`._", ""]
    for title, items in sections:
        lines.append(f"## {title}\n")
        for lvl, msg in items:
            counts[lvl] += 1
            badge = {"CRITICAL": "🔴", "WARN": "🟠", "INFO": "ℹ️", "OK": "✅"}[lvl]
            lines.append(f"- {badge} **{lvl}** — {msg}")
        lines.append("")
    lines.append(f"## Paid call-site inventory ({len(inv)} hits)\n")
    by = {}
    for label, f, l in inv:
        by.setdefault(label, []).append(f"`{f}:{l}`")
    for label, locs in sorted(by.items()):
        lines.append(f"- **{label}** ({len(locs)}): " + ", ".join(locs[:12]) + (" …" if len(locs) > 12 else ""))
    lines.append(f"\n**Totals:** {counts[CRIT]} critical · {counts[WARN]} warnings · {counts[OK]} ok · {counts[INFO]} info\n")

    rec = root / MAIN_REPO / "Records"
    rec.mkdir(exist_ok=True)
    out_path = rec / f"COST_SWEEP_{today}.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    if not args.quiet:
        print("\n".join(lines))
    print(f"\nreport → {out_path}")
    sys.exit(2 if counts[CRIT] else (1 if counts[WARN] else 0))

if __name__ == "__main__":
    main()
